"""
delaware Timesheet Automator
"""
import json, os, sys, webbrowser, http.server, threading
import urllib.parse, urllib.request, urllib.error
import datetime, time
from pathlib import Path

CONFIG_FILE   = Path(__file__).parent / "config.json"
PROJECTS_FILE = Path(__file__).parent / "projects.xlsx"
TIME_BASE     = "https://time.delaware.pro/api/v1/timesheets"
GEMINI_API = "https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-flash:generateContent"

def load_config():
    if CONFIG_FILE.exists():
        with open(CONFIG_FILE, encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_config(cfg):
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(cfg, f, indent=2, ensure_ascii=False)

def load_projects_from_excel():
    if not PROJECTS_FILE.exists():
        return None, "projects.xlsx not found"
    try:
        import openpyxl
        wb = openpyxl.load_workbook(PROJECTS_FILE, data_only=True)
        ws = wb["Projects"]
        headers = [c.value for c in ws[1]]
        col = {h: i for i, h in enumerate(headers) if h}
        projects = []
        today = datetime.date.today()
        for row in ws.iter_rows(min_row=2, values_only=True):
            code = row[col.get("Project Code", 0)]
            if not code:
                continue
            active = str(row[col.get("Active", 9)] or "").strip().lower()
            if active != "yes":
                continue
            start = row[col.get("Start Date", 7)]
            end   = row[col.get("End Date", 8)]
            if start and hasattr(start, "date"): start = start.date()
            if end   and hasattr(end,   "date"): end   = end.date()
            if isinstance(start, datetime.date) and start > today: continue
            if isinstance(end,   datetime.date) and end   < today: continue
            projects.append({
                "projectCode":         str(code).strip(),
                "projectTaskCode":     str(row[col.get("Task Code", 1)] or "").strip(),
                "projectTaskItemCode": "M001",
                "title":               str(row[col.get("Client Name", 2)] or "").strip(),
                "label":               str(row[col.get("Project Label", 3)] or "").strip(),
                "lane1":               str(row[col.get("Project Label", 3)] or "").strip(),
                "lane3":               "",
                "clientName":          str(row[col.get("Client Name", 2)] or "").strip(),
                "ptype":               str(row[col.get("Type", 4)] or "").strip(),
                "teamMembers":         str(row[col.get("Team Members", 5)] or "").strip(),
                "tags":                str(row[col.get("Tags (comma separated)", 6)] or "").lower().strip(),
                "codeCategory":        str(row[col.get("Type", 4)] or "").strip(),
            })
        return projects, None
    except Exception as e:
        return None, str(e)

def load_corrections():
    if not PROJECTS_FILE.exists():
        return {}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(PROJECTS_FILE, data_only=True)
        if "Corrections Log" not in wb.sheetnames:
            return {}
        ws = wb["Corrections Log"]
        corrections = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                corrections[str(row[0]).lower().strip()] = {
                    "projectCode": str(row[1] or ""),
                    "projectTaskCode": str(row[2] or ""),
                    "workDescription": str(row[3] or ""),
                }
        return corrections
    except Exception:
        return {}

def save_correction(subject, project_code, task_code, description):
    if not PROJECTS_FILE.exists():
        return
    try:
        import openpyxl
        wb = openpyxl.load_workbook(PROJECTS_FILE)
        if "Corrections Log" not in wb.sheetnames:
            return
        ws = wb["Corrections Log"]
        key = subject.lower().strip()
        for row in ws.iter_rows(min_row=2):
            if row[0].value and str(row[0].value).lower().strip() == key:
                row[1].value = project_code
                row[2].value = task_code
                row[3].value = description
                row[4].value = datetime.date.today().isoformat()
                wb.save(PROJECTS_FILE)
                return
        ws.append([subject, project_code, task_code, description,
                   datetime.date.today().isoformat()])
        wb.save(PROJECTS_FILE)
    except Exception:
        pass

def _parse_applescript_events(output):
    events = []
    for line in output.strip().splitlines():
        parts = line.split("|||")
        if len(parts) < 4:
            continue
        subject, start_fmt, end_fmt, dur_str = parts[0], parts[1], parts[2], parts[3]
        try:
            duration_hours = round(float(dur_str) / 60 * 4) / 4
            if duration_hours >= 0.25:
                events.append({
                    "subject":         subject or "(No subject)",
                    "_duration_hours": duration_hours,
                    "_start_fmt":      start_fmt,
                    "_end_fmt":        end_fmt,
                })
        except ValueError:
            continue
    return events

def _read_outlook_mac(target_date):
    import subprocess
    y, mo, d = target_date.year, target_date.month, target_date.day
    script = f"""
tell application "Microsoft Outlook"
    set startDate to current date
    set year of startDate to {y}
    set month of startDate to {mo}
    set day of startDate to {d}
    set hours of startDate to 0
    set minutes of startDate to 0
    set seconds of startDate to 0
    set endDate to startDate + 86399
    set output to ""
    repeat with e in (every calendar event whose start time >= startDate and start time <= endDate)
        if all day event of e is false then
            set eStart to start time of e
            set eEnd to end time of e
            set eDur to (eEnd - eStart) / 60
            if eDur >= 15 then
                set sH to text -2 thru -1 of ("0" & (hours of eStart as string))
                set sM to text -2 thru -1 of ("0" & (minutes of eStart as string))
                set eH to text -2 thru -1 of ("0" & (hours of eEnd as string))
                set eM to text -2 thru -1 of ("0" & (minutes of eEnd as string))
                set output to output & (subject of e) & "|||" & sH & ":" & sM & "|||" & eH & ":" & eM & "|||" & eDur & "\\n"
            end if
        end if
    end repeat
    return output
end tell
"""
    try:
        r = subprocess.run(["osascript", "-e", script], capture_output=True, text=True, timeout=30)
        if r.returncode != 0:
            return None, r.stderr.strip()
        return _parse_applescript_events(r.stdout), None
    except Exception as e:
        return None, str(e)

def _read_calendar_app_mac(target_date):
    import subprocess
    y, mo, d = target_date.year, target_date.month, target_date.day
    script = f"""
tell application "Calendar"
    set startDate to current date
    set year of startDate to {y}
    set month of startDate to {mo}
    set day of startDate to {d}
    set hours of startDate to 0
    set minutes of startDate to 0
    set seconds of startDate to 0
    set endDate to startDate + 86399
    set output to ""
    repeat with c in every calendar
        repeat with e in (every event of c whose start date >= startDate and start date <= endDate)
            if allday event of e is false then
                set eStart to start date of e
                set eEnd to end date of e
                set eDur to (eEnd - eStart) / 60
                if eDur >= 15 then
                    set sH to text -2 thru -1 of ("0" & (hours of eStart as string))
                    set sM to text -2 thru -1 of ("0" & (minutes of eStart as string))
                    set eH to text -2 thru -1 of ("0" & (hours of eEnd as string))
                    set eM to text -2 thru -1 of ("0" & (minutes of eEnd as string))
                    set output to output & (summary of e) & "|||" & sH & ":" & sM & "|||" & eH & ":" & eM & "|||" & eDur & "\\n"
                end if
            end if
        end repeat
    end repeat
    return output
end tell
"""
    try:
        r = subprocess.run(["osascript", "-e", script], capture_output=True, text=True, timeout=30)
        if r.returncode != 0:
            return None, r.stderr.strip()
        return _parse_applescript_events(r.stdout), None
    except Exception as e:
        return None, str(e)

def read_outlook_calendar(target_date):
    import platform
    if platform.system() == "Darwin":
        events, err = _read_outlook_mac(target_date)
        if events is not None:
            return events, None
        print(f"        Outlook for Mac unavailable ({err}), trying Calendar.app...")
        events, err = _read_calendar_app_mac(target_date)
        return events, err

    # Windows fallback
    try:
        import win32com.client
    except ImportError:
        return None, "pywin32 not installed"
    try:
        outlook  = win32com.client.Dispatch("Outlook.Application")
        ns       = outlook.GetNamespace("MAPI")
        calendar = ns.GetDefaultFolder(9)
        items    = calendar.Items
        items.IncludeRecurrences = True
        items.Sort("[Start]")
        sf = target_date.strftime("%m/%d/%Y 00:00 AM")
        ef = target_date.strftime("%m/%d/%Y 11:59 PM")
        items = items.Restrict(f"[Start] >= '{sf}' AND [Start] <= '{ef}'")
        events = []
        for item in items:
            try:
                if item.AllDayEvent or item.Duration < 15:
                    continue
                s  = item.Start
                en = item.End
                sd = datetime.datetime(s.year, s.month, s.day, s.hour, s.minute)
                ed = datetime.datetime(en.year, en.month, en.day, en.hour, en.minute)
                events.append({
                    "subject":         item.Subject or "(No subject)",
                    "_duration_hours": round(item.Duration / 60 * 4) / 4,
                    "_start_fmt":      sd.strftime("%H:%M"),
                    "_end_fmt":        ed.strftime("%H:%M"),
                })
            except Exception:
                continue
        return events, None
    except Exception as e:
        return None, f"outlook_error:{e}"

def fetch_project_codes_api(token, user_id, date_str):
    url = f"{TIME_BASE}/staffedprojectcodes?date={date_str}&user="
    print(f"        Fetching: {url}")
    req = urllib.request.Request(url, headers={
        "Authorization": f"Bearer {token}",
        "Accept": "application/json",
        "Origin": "https://time.delaware.pro",
        "Referer": "https://time.delaware.pro/",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/145.0.0.0 Safari/537.36 Edg/145.0.0.0",
    })
    try:
        raw = urllib.request.urlopen(req, timeout=10).read()
        data = json.loads(raw)
        print(f"        Response keys: {list(data.keys()) if isinstance(data, dict) else type(data).__name__}")
        if isinstance(data, dict) and "codes" in data:
            return data["codes"]
        if isinstance(data, list):
            return data
        return data.get("data", [])
    except urllib.error.HTTPError as e:
        body = e.read().decode("utf-8", errors="ignore")
        print(f"        API error HTTP {e.code}: {body[:200]}")
        return []
    except Exception as e:
        print(f"        API error: {e}")
        return []

def ai_map_events(events, projects, corrections, target_date, gemini_key):
    date_str = target_date.strftime("%Y-%m-%d")
    pc_lines = "\n".join(
        "- " + p["projectCode"] + " | " + p["projectTaskCode"] + " | " +
        (p.get("title") or p.get("label") or p.get("clientName") or "") + " | tags: " + p.get("tags","")
        for p in projects
    ) if projects else "No projects loaded"

    pre_mapped = {}
    for e in events:
        key = e["subject"].lower().strip()
        for pattern, correction in corrections.items():
            if pattern in key or key in pattern:
                pre_mapped[e["subject"]] = correction
                break

    ev_lines = "\n".join(
        "- \"" + e["subject"] + "\" | " + e["_start_fmt"] + "-" + e["_end_fmt"] +
        " | " + str(e["_duration_hours"]) + "h" +
        (" [CORRECTION: use " + pre_mapped[e["subject"]]["projectCode"] + "]"
         if e["subject"] in pre_mapped else "")
        for e in events
    )

    prompt = (
        "You are a timesheet assistant for a SAP consultant at delaware consulting in Belgium.\n\n"
        "Available projects (projectCode | taskCode | label | client | tags):\n"
        + pc_lines + "\n\n"
        "Calendar events for " + date_str + ":\n"
        + ev_lines + "\n\n"
        "Entries marked [CORRECTION] MUST use the specified project code.\n\n"
        "Respond ONLY with a valid JSON array, no markdown:\n"
        '[{"subject":"exact event subject","activityDate":"' + date_str + '",'
        '"duration":1.5,"projectCode":"CODE","projectTaskCode":"CODE.1.1",'
        '"projectTaskItemCode":"M001","workDescription":"Short description max 60 chars",'
        '"confidence":"high","reason":"One sentence why"}]\n\n'
        "Rules: duration=event hours rounded to 0.25, projectTaskItemCode always M001, "
        "confidence=high/medium/low, internal meetings use Internal type project, "
        "workDescription in professional English MAX 50 chars, strictly enforced."
    )

    data = json.dumps({
        "contents": [{"parts": [{"text": prompt}]}]
    }).encode()

    url = GEMINI_API + "?key=" + gemini_key
    req = urllib.request.Request(url, data=data, headers={"Content-Type": "application/json"})
    try:
        raw = urllib.request.urlopen(req, timeout=30).read()
    except urllib.error.HTTPError as e:
        raise Exception("HTTP " + str(e.code) + ": " + e.read().decode())
    resp = json.loads(raw)
    text = resp["candidates"][0]["content"]["parts"][0]["text"].replace("```json", "").replace("```", "").strip()
    return json.loads(text)

def build_review_html(draft_rows, title, projects, user_id, time_token):
    def _opt_label(p):
        t = p.get("title") or p.get("label") or p.get("clientName") or p.get("lane1") or ""
        l3 = p.get("lane3") or ""
        return (t + (" — " + l3 if l3 else ""))[:70]

    pc_options = "".join(
        '<option value="' + p.get("projectCode","") +
        '" data-task="' + p.get("projectTaskCode","") +
        '" data-item="' + p.get("projectTaskItemCode","") + '">' +
        p.get("projectCode","") + " — " + _opt_label(p) + "</option>"
        for p in projects
    ) if projects else '<option value="">No projects loaded</option>'

    rows_html = ""
    for i, r in enumerate(draft_rows):
        cc  = {"high": "conf-high", "medium": "conf-med", "low": "conf-low"}.get(
              r.get("confidence", "low"), "conf-low")
        sel = pc_options.replace(
            'value="' + r.get("projectCode","") + '"',
            'value="' + r.get("projectCode","") + '" selected', 1)
        desc = r.get("workDescription", "").replace('"', "&quot;")
        rows_html += (
            '<tr id="row-' + str(i) + '">'
            '<td class="ce">' + r.get("subject","") + '</td>'
            '<td><select id="pc-' + str(i) + '" onchange="uc(' + str(i) + ',this)">' + sel + '</select></td>'
            '<td><input type="text" id="desc-' + str(i) + '" value="' + desc + '" maxlength="50"/></td>'
            '<td><input type="number" id="dur-' + str(i) + '" value="' + str(r.get("duration",1)) +
            '" min="0.25" max="10" step="0.25" style="width:60px"/></td>'
            '<td><span class="badge ' + cc + '">' + r.get("confidence","?") + '</span>'
            '<div class="rsn">' + r.get("reason","") + '</div></td>'
            '<td><button onclick="rr(' + str(i) + ')" class="db">\xd7</button></td>'
            '</tr>'
        )

    rows_json    = json.dumps(draft_rows, ensure_ascii=False)
    token_style  = "display:none" if time_token else ""
    token_filled = time_token

    css = """*{box-sizing:border-box;margin:0;padding:0}
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#f5f5f3;color:#1a1a1a;font-size:14px}
.tb{background:#1a1a1a;color:#fff;padding:14px 24px;display:flex;align-items:center;gap:12px}
.tb h1{font-size:16px;font-weight:500;flex:1}.tb .ss{font-size:12px;color:#aaa}
.wrap{max-width:1100px;margin:20px auto;padding:0 16px}
.card{background:#fff;border:1px solid #e5e5e3;border-radius:10px;overflow:hidden;margin-bottom:12px}
.ch{padding:12px 20px;border-bottom:1px solid #e5e5e3;display:flex;align-items:center}
.ch h2{font-size:14px;font-weight:500;flex:1}
.stats{display:flex;gap:20px}
.sv{font-size:20px;font-weight:500}.sl{font-size:11px;color:#888;text-align:right}
.trow{padding:10px 20px;background:#fffbee;border-bottom:1px solid #f0e8c8;display:flex;align-items:center;gap:10px}
.trow label{font-size:12px;font-weight:500;white-space:nowrap}
.trow input{flex:1;border:1px solid #ddd;border-radius:6px;padding:5px 10px;font-size:12px;font-family:monospace}
.trow .th{font-size:11px;color:#888}
table{width:100%;border-collapse:collapse}
th{font-size:11px;font-weight:500;color:#888;text-align:left;padding:7px 10px;border-bottom:1px solid #e5e5e3}
td{padding:7px 10px;border-bottom:1px solid #f0f0ee;vertical-align:middle}
tr:last-child td{border-bottom:none}
tr:hover td{background:#fafaf8}
.ce{font-size:12px;color:#555;max-width:150px}
select{width:100%;min-width:190px;border:1px solid #ddd;border-radius:6px;padding:5px 7px;font-size:12px}
input[type=text]{width:100%;min-width:190px;border:1px solid #ddd;border-radius:6px;padding:5px 7px;font-size:12px}
input:focus,select:focus{outline:none;border-color:#1a1a1a}
.badge{display:inline-block;font-size:10px;padding:2px 6px;border-radius:99px;font-weight:500}
.conf-high{background:#d4edda;color:#155724}.conf-med{background:#fff3cd;color:#856404}.conf-low{background:#f8d7da;color:#721c24}
.rsn{font-size:10px;color:#aaa;margin-top:2px;max-width:150px;line-height:1.3}
.db{border:none;background:none;color:#ccc;font-size:18px;cursor:pointer;padding:2px 5px}
.db:hover{color:#e55}
.ac{padding:14px 20px;display:flex;gap:10px;align-items:center;flex-wrap:wrap}
.btn{padding:8px 18px;border-radius:7px;font-size:13px;font-weight:500;cursor:pointer;border:1px solid #ddd;background:#fff;color:#1a1a1a}
.btn:hover{background:#f0f0ee}
.bp{background:#1a1a1a;color:#fff;border-color:transparent}
.bp:hover{background:#333}
#st{font-size:13px;margin-left:auto}
.ok{color:#155724}.err{color:#721c24}.inf{color:#0c5460}.wrn{color:#856404}
.hw{color:#856404}.hok{color:#155724}"""

    first_date = draft_rows[0].get("activityDate", "") if draft_rows else ""

    js = (
        "var rows=" + rows_json + ";"
        "var ds='" + first_date + "';"
        "var uid='" + user_id + "';"
        "var preToken='" + token_filled + "';"
        "function us(){"
        "  var t=0,durs=document.querySelectorAll('[id^=\"dur-\"]');"
        "  for(var i=0;i<durs.length;i++){t+=parseFloat(durs[i].value)||0;}"
        "  var el=document.getElementById('th2');el.textContent=t.toFixed(2);"
        "  el.className='sv '+(t<4||t>10?'hw':'hok');"
        "  document.getElementById('ec').textContent=document.querySelectorAll('#tb2 tr').length;"
        "}"
        "function uc(i,s){"
        "  if(rows[i]){rows[i].projectCode=s.value;"
        "  rows[i].projectTaskCode=s.options[s.selectedIndex].dataset.task||s.value+'.1.1';}"
        "}"
        "function rr(i){var r=document.getElementById('row-'+i);if(r){r.remove();us();}}"
        "function ar(){"
        "  var i=rows.length;"
        "  rows.push({subject:'Manual',projectCode:'',projectTaskCode:'',projectTaskItemCode:'M001',workDescription:'',duration:1});"
        "  var tb=document.getElementById('tb2'),tr=document.createElement('tr');tr.id='row-'+i;"
        "  var pcOpts=document.querySelector('[id^=\"pc-\"]');"
        "  var optHtml=pcOpts?pcOpts.innerHTML:'<option value=\"\">Select...</option>';"
        "  tr.innerHTML='<td class=\"ce\"><input type=\"text\" style=\"width:130px;border:1px solid #ddd;border-radius:6px;padding:4px 7px;font-size:12px\" value=\"Manual entry\"/></td>'"
        "    +'<td><select id=\"pc-'+i+'\" onchange=\"uc('+i+',this)\">'+optHtml+'</select></td>'"
        "    +'<td><input type=\"text\" id=\"desc-'+i+'\" value=\"\" maxlength=\"80\"/></td>'"
        "    +'<td><input type=\"number\" id=\"dur-'+i+'\" value=\"1\" min=\"0.25\" max=\"10\" step=\"0.25\" style=\"width:60px\"/></td>'"
        "    +'<td></td><td><button onclick=\"rr('+i+')\" class=\"db\">\xd7</button></td>';"
        "  tb.appendChild(tr);us();"
        "}"
        "function ss2(m,c){var e=document.getElementById('st');e.textContent=m;e.className=c;}"
        "function getToken(){var inp=document.getElementById('tok');return (inp?inp.value.trim():'')||preToken;}"
        "async function sub(){"
        "  var t=getToken();"
        "  if(!t){ss2('Paste your bearer token first.','wrn');return;}"
        "  var en=[],corr=[];"
        "  var sels=document.querySelectorAll('[id^=\"pc-\"]');"
        "  for(var si=0;si<sels.length;si++){"
        "    var s=sels[si],i=parseInt(s.id.replace('pc-',''));"
        "    var d=document.getElementById('desc-'+i),h=document.getElementById('dur-'+i);"
        "    if(!s.value)continue;"
        "    var orig=rows[i]?rows[i].projectCode:'';"
        "    var desc=d?d.value:'';"
        "    var pc=s.value;"
        "    var ptc=rows[i]&&rows[i].projectTaskCode||s.options[s.selectedIndex].dataset.task||pc+'.1.1';"
        "    var ad=rows[i]?rows[i].activityDate:'';"
        "    var ptic=s.options[s.selectedIndex].dataset.item!==undefined?s.options[s.selectedIndex].dataset.item:'';\n"
        "    en.push({projectCode:pc,projectTaskCode:ptc,projectTaskItemCode:ptic,"
        "      activityDate:ad,duration:parseFloat(h?h.value:1),workDescription:desc,userId:uid});"
        "    if(orig&&orig!==pc&&rows[i]){"
        "      corr.push({subject:rows[i].subject,projectCode:pc,taskCode:ptc,description:desc});}"
        "  }"
        "  if(!en.length){ss2('No entries to submit.','wrn');return;}"
        "  ss2('Submitting '+en.length+' entries...','inf');"
        "  try{"
        "    var r=await fetch('/submit',{method:'POST',"
        "      headers:{'Content-Type':'application/json','X-Token':t},"
        "      body:JSON.stringify({entries:en,corrections:corr})});"
        "    var data=await r.json();"
        "    if(data.ok){"
        "      ss2('✓ '+data.count+' submitted'+(corr.length?' + '+corr.length+' correction(s) saved':'')+'!','ok');"
        "      document.getElementById('ss').textContent='Submitted ✓';"
        "    }else ss2('Error: '+(data.error||'?'),'err');"
        "  }catch(e){ss2('Failed: '+e.message,'err');}"
        "}"
        "us();"
        "document.querySelectorAll('[id^=\"pc-\"]').forEach(function(sel,idx){"
        "  if(rows[idx]&&rows[idx].projectCode){"
        "    sel.value=rows[idx].projectCode;}"
        "});"
    )

    # Fix the activityDate js var which has a Python expression bug above
    # Just use the first row's date or empty string safely
    first_date = draft_rows[0].get("activityDate", "") if draft_rows else ""
    js = js.replace(
        "var ds='" + rows_json + "' and " + repr(draft_rows[0].get("activityDate","") if draft_rows else "") + "if " + repr(bool(draft_rows)) + " else '' + "+"'';",
        "var ds='" + first_date + "';"
    )

    return (
        "<!DOCTYPE html><html lang='en'><head><meta charset='UTF-8'>"
        "<title>Timesheet \u2014 " + title + "</title>"
        "<style>" + css + "</style></head><body>"
        "<div class='tb'>"
        "<svg width='18' height='18' viewBox='0 0 20 20' fill='none'>"
        "<rect x='3' y='2' width='14' height='16' rx='2' fill='none' stroke='white' stroke-width='1.5'/>"
        "<path d='M7 7h6M7 10h6M7 13h4' stroke='white' stroke-width='1.5' stroke-linecap='round'/>"
        "</svg>"
        "<h1>Timesheet draft \u2014 " + title + "</h1>"
        "<span class='ss' id='ss'>Not submitted yet</span>"
        "</div>"
        "<div class='wrap'><div class='card'>"
        "<div class='trow' style='" + token_style + "'>"
        "<label>Bearer token:</label>"
        "<input type='password' id='tok' value='" + token_filled + "' "
        "placeholder='Paste bearer token from time.delaware.pro network tab'/>"
        "<span class='th'>Paste once per session</span>"
        "</div>"
        "<div class='ch'><h2>Entries</h2>"
        "<div class='stats'>"
        "<div><div class='sv' id='th2'>0.0</div><div class='sl'>hours</div></div>"
        "<div style='margin-left:16px'><div class='sv' id='ec'>0</div><div class='sl'>entries</div></div>"
        "</div></div>"
        "<table><thead><tr>"
        "<th>Calendar event</th><th>Project code</th><th>Work description</th>"
        "<th>Hours</th><th>Confidence</th><th></th>"
        "</tr></thead>"
        "<tbody id='tb2'>" + rows_html + "</tbody></table>"
        "<div class='ac'>"
        "<button class='btn bp' onclick='sub()'>Submit to timesheet \u2192</button>"
        "<button class='btn' onclick='ar()'>+ add row</button>"
        "<span id='st'></span>"
        "</div></div></div>"
        "<script>" + js + "</script>"
        "</body></html>"
    )

class Handler(http.server.BaseHTTPRequestHandler):
    html = ""; user_id = ""
    def log_message(self, *a): pass
    def do_GET(self):
        if self.path == "/":
            self.send_response(200)
            self.send_header("Content-Type", "text/html;charset=utf-8")
            self.end_headers()
            self.wfile.write(Handler.html.encode())
    def do_POST(self):
        if self.path == "/submit":
            n    = int(self.headers.get("Content-Length", 0))
            tok  = self.headers.get("X-Token", "")
            body = json.loads(self.rfile.read(n))
            entries = body.get("entries", [])

            ok_n, errs = 0, []
            for e in entries:
                r = submit_entry(tok, Handler.user_id, e)
                if r["ok"]: ok_n += 1
                else: errs.append(r.get("error", "?"))

            for c in body.get("corrections", []):
                save_correction(c["subject"], c["projectCode"], c["taskCode"], c["description"])

            resp = {"ok": not errs, "count": ok_n}
            if errs: resp["error"] = "; ".join(errs[:3])

            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.end_headers()
            self.wfile.write(json.dumps(resp).encode())

def start_server(html, user_id, port=8421):
    Handler.html = html; Handler.user_id = user_id
    srv = http.server.HTTPServer(("localhost", port), Handler)
    threading.Thread(target=srv.serve_forever, daemon=True).start()
    return srv

def submit_entry(token, user_id, entry):
    payload = json.dumps({"data": {
        "userId": "",
        "projectCode": entry["projectCode"],
        "projectTaskCode": entry["projectTaskCode"],
        "projectTaskItemCode": entry.get("projectTaskItemCode", "M001"),
        "activityDate": entry["activityDate"],
        "duration": float(str(entry["duration"]).replace(",", ".")),
        "workDescription": str(entry.get("workDescription", ""))[:50],
        "externalReferences": [],
        "properties": []
    }}).encode()
    url = f"{TIME_BASE}/timeentry?user={urllib.parse.quote(user_id.upper())}"
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
        "Accept": "application/json, text/plain, */*",
        "Origin": "https://time.delaware.pro",
        "Referer": "https://time.delaware.pro/",
        "X-Requested-With": "XMLHttpRequest",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/145.0.0.0 Safari/537.36 Edg/145.0.0.0",
    }
    req = urllib.request.Request(url, data=payload, headers=headers)
    req.get_method = lambda: "POST"
    try:
        urllib.request.urlopen(req, timeout=10)
        return {"ok": True}
    except urllib.error.HTTPError as e:
        return {"ok": False, "error": f"HTTP {e.code}"}
    except Exception as e:
        return {"ok": False, "error": str(e)}

def process_day(target_date, projects, corrections, cfg):
    print(f"\n  ── {target_date.strftime('%A, %d %B %Y')} ──")
    events, err = read_outlook_calendar(target_date)
    if err or not events:
        print(f"        No events found — skipping")
        return None
    print(f"        {len(events)} events found")
    for e in events:
        print(f"          {e['_start_fmt']}-{e['_end_fmt']}  {e['subject'][:55]}")
    print(f"        AI mapping...")
    try:
        draft = ai_map_events(events, projects, corrections, target_date, cfg["gemini_key"])
        print(f"        {len(draft)} entries drafted")
        return draft
    except Exception as ex:
        print(f"        AI failed ({ex}) — basic draft")
        return [{"subject": e["subject"], "activityDate": str(target_date),
                 "duration": e["_duration_hours"], "projectCode": "",
                 "projectTaskCode": "", "projectTaskItemCode": "M001",
                 "workDescription": e["subject"][:60],
                 "confidence": "low", "reason": "AI unavailable"}
                for e in events]

def main():
    print("\n╔══════════════════════════════════════╗")
    print("║   delaware Timesheet Automator       ║")
    print("╚══════════════════════════════════════╝\n")

    cfg     = load_config()
    user_id = cfg.get("user_id", "")

    if not cfg.get("gemini_key") or "PASTE" in cfg.get("gemini_key", ""):
        print("  Set your Gemini API key in config.json\n")
        input("  Press Enter to exit..."); sys.exit(1)
    if not user_id or "PASTE" in user_id:
        print("  Set your user_id in config.json\n")
        input("  Press Enter to exit..."); sys.exit(1)

    # Load projects from Excel (for tags/AI mapping)
    print("  Loading projects from projects.xlsx...")
    excel_projects, err = load_projects_from_excel()
    if err:
        print(f"  Warning: {err} — using config.json fallback")
        excel_projects = [
            {"projectCode": k, "projectTaskCode": v["taskCode"],
             "projectTaskItemCode": "M001",
             "title": v["label"], "label": v["label"],
             "lane1": v["label"], "lane3": "",
             "clientName": "", "ptype": "",
             "teamMembers": "", "tags": ",".join(v.get("keywords", []))}
            for k, v in cfg.get("project_codes", {}).items()
        ]
    else:
        print(f"  {len(excel_projects)} active projects loaded from Excel")

    corrections = load_corrections()
    if corrections:
        print(f"  {len(corrections)} learned corrections loaded")

    # Bearer token
    print("\n  Open time.delaware.pro in Edge, then:")
    print("  F12 -> Network -> Fetch/XHR -> click any timeentry request")
    print("  -> Headers -> copy value after 'Bearer '\n")
    time_token = input("  Paste bearer token: ").strip()
    print()

    # Fetch live project codes from API for dropdown
    today = datetime.date.today()
    fetch_date = str(today - datetime.timedelta(days=1) if "--today" not in sys.argv else today)
    api_projects = []
    if time_token:
        print("  Fetching live project codes from time.delaware.pro...")
        api_projects = fetch_project_codes_api(time_token, user_id, fetch_date)
        if api_projects:
            print(f"  {len(api_projects)} project codes loaded from API:")
            for p in api_projects[:8]:
                print(f"    {p.get('projectCode','')} | {p.get('projectTaskCode','')} | {p.get('projectTaskItemCode','')} | {(p.get('title','') or '')} — {(p.get('lane3','') or '')}"[:80])
            if len(api_projects) > 8:
                print(f"    ... and {len(api_projects)-8} more")
        else:
            print("  Could not load from API — using Excel projects for dropdown")

    # Use API codes for dropdown, Excel codes for AI tag matching
    dropdown_projects = api_projects if api_projects else excel_projects
    ai_projects = excel_projects  # Excel has tags, API doesn't

    # Date selection
    if "--today" in sys.argv:
        dates_to_process = [today]
    else:
        candidates = []
        d = today - datetime.timedelta(days=1)
        while len(candidates) < 7:
            if d.weekday() < 5:
                candidates.append(d)
            d -= datetime.timedelta(days=1)

        print("\n  Which days to process?")
        print("  (Just press Enter for yesterday only)\n")
        for i, day in enumerate(candidates):
            tag = " <- yesterday" if i == 0 else ""
            print(f"  [{i+1}] {day.strftime('%A, %d %B %Y')}{tag}")
        print()
        choice = input("  Enter number(s), e.g. 1 or 1,2,3: ").strip()

        if not choice:
            dates_to_process = [candidates[0]]
        else:
            try:
                indices = [int(x.strip()) - 1
                           for x in choice.replace(",", " ").split() if x.strip()]
                dates_to_process = [candidates[i] for i in indices if 0 <= i < len(candidates)]
            except Exception:
                dates_to_process = [candidates[0]]

    print(f"\n  Processing {len(dates_to_process)} day(s)...")

    all_drafts = {}
    for d in dates_to_process:
        draft = process_day(d, ai_projects, corrections, cfg)
        if draft is not None:
            all_drafts[d] = draft

    if not all_drafts:
        print("\n  No events found for any selected day.")
        input("  Press Enter to exit...")
        return

    combined = []
    for draft in all_drafts.values():
        combined.extend(draft)

    if len(all_drafts) == 1:
        d = list(all_drafts.keys())[0]
        title = d.strftime("%A, %d %B %Y")
    else:
        title = f"{len(all_drafts)} days"

    html   = build_review_html(combined, title, dropdown_projects, user_id, time_token)
    server = start_server(html, user_id)

    print(f"\n  Opening review page — {len(combined)} entries")
    print(f"  Review, edit if needed, then click Submit.\n")
    print(f"  Press Ctrl+C when done.\n")

    webbrowser.open("http://localhost:8421")
    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        server.shutdown()
        print("\n  Done!\n")

if __name__ == "__main__":
    main()
